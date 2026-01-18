import os
from typing import List, Tuple, TYPE_CHECKING
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from dance_loop_gen.core.report_models import ReportData, ReportRow
from dance_loop_gen.utils.logger import setup_logger

logger = setup_logger()

class ReportService:
    """Service for generating Excel reports with embedded visuals."""

    # Constants for layout
    HEADER_ROW = 4
    ROW_HEIGHT = 130
    TARGET_IMG_HEIGHT = 160

    # Styles
    _THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    _CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    _WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")

    def generate_report(self, data: ReportData, output_path: str) -> str:
        """Generates an Excel report from ReportData."""
        logger.info(f"Generating Excel report for: {data.title}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Production Schedule"

        self._write_title(ws, data.title)
        self._write_metadata(ws, data)
        self._setup_table_headers(ws)
        self._write_data_rows(ws, data.rows)

        wb.save(output_path)
        logger.info(f"Report saved to: {output_path}")
        return output_path

    def _write_title(self, ws: Worksheet, title: str) -> None:
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = f"VIDEO PLAN: {title.upper()}"
        cell.font = Font(size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = self._CENTER_ALIGN

    def _write_metadata(self, ws: Worksheet, data: ReportData) -> None:
        ws.merge_cells('A2:G2')
        cell = ws['A2']
        cell.value = f"Run ID: {data.run_id} | Generated: {data.generated_at}"
        cell.font = Font(italic=True, color="555555")
        cell.alignment = self._CENTER_ALIGN

    def _setup_table_headers(self, ws: Worksheet) -> None:
        headers = [
            ("Scene", 8), ("Keyframe", 25), ("Action", 40),
            ("Audio", 25), ("Start Pose", 25), ("End Pose", 25), ("Notes", 20)
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="36454F", end_color="36454F", fill_type="solid")

        for col_num, (text, width) in enumerate(headers, 1):
            cell = ws.cell(row=self.HEADER_ROW, column=col_num, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = self._CENTER_ALIGN
            cell.border = self._THIN_BORDER
            ws.column_dimensions[chr(64 + col_num)].width = width

    def _write_data_rows(self, ws: Worksheet, rows: List[ReportRow]) -> None:
        for i, row_data in enumerate(rows, start=self.HEADER_ROW + 1):
            ws.row_dimensions[i].height = self.ROW_HEIGHT
            self._write_single_row(ws, i, row_data)

    def _write_single_row(self, ws: Worksheet, row_idx: int, data: ReportRow) -> None:
        # Scene Number
        c1 = ws.cell(row=row_idx, column=1, value=data.scene_number)
        c1.alignment = self._CENTER_ALIGN
        c1.font = Font(bold=True, size=12)
        c1.border = self._THIN_BORDER

        # Image
        c2 = ws.cell(row=row_idx, column=2)
        c2.border = self._THIN_BORDER
        self._insert_image(ws, row_idx, data.keyframe_path, c2)

        # Text Columns (3-7)
        values = [data.action_description, data.audio_prompt,
                  data.start_pose, data.end_pose, data.notes]

        for col_offset, val in enumerate(values):
            cell = ws.cell(row=row_idx, column=3 + col_offset, value=val)
            cell.alignment = self._WRAP_ALIGN
            cell.border = self._THIN_BORDER

    def _insert_image(self, ws: Worksheet, row_idx: int, path: str, cell) -> None:
        if not path or not os.path.exists(path):
            cell.value = "[No Image]"
            cell.alignment = self._CENTER_ALIGN
            return

        try:
            img = XLImage(path)
            scale = self.TARGET_IMG_HEIGHT / img.height
            img.height = self.TARGET_IMG_HEIGHT
            img.width = int(img.width * scale)

            ws.add_image(img, f"B{row_idx}")
        except Exception as e:
            logger.error(f"Failed to add image {path}: {e}")
            cell.value = "[Image Error]"
            cell.alignment = self._CENTER_ALIGN

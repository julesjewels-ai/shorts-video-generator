import os
import unittest
from datetime import datetime
from PIL import Image
from openpyxl import load_workbook
from dance_loop_gen.services.report_service import ReportService
from dance_loop_gen.core.report_models import ReportData, ReportRow

class TestReportService(unittest.TestCase):
    def setUp(self):
        self.output_file = "test_report.xlsx"
        self.test_image_file = "test_image.png"

        # Create a dummy image
        img = Image.new('RGB', (100, 100), color = 'red')
        img.save(self.test_image_file)

        self.service = ReportService()

    def tearDown(self):
        if os.path.exists(self.output_file):
            os.remove(self.output_file)
        if os.path.exists(self.test_image_file):
            os.remove(self.test_image_file)

    def test_generate_report(self):
        # Arrange
        data = ReportData(
            title="Test Plan",
            generated_at="2023-01-01 12:00",
            run_id="run_123",
            rows=[
                ReportRow(
                    scene_number=1,
                    keyframe_path=self.test_image_file,
                    action_description="Action 1",
                    audio_prompt="Audio 1",
                    start_pose="Pose Start",
                    end_pose="Pose End",
                    notes="Note 1"
                ),
                ReportRow(
                    scene_number=2,
                    keyframe_path="non_existent_image.png",
                    action_description="Action 2",
                    audio_prompt="Audio 2",
                    start_pose="Pose Start 2",
                    end_pose="Pose End 2",
                    notes="Note 2"
                )
            ]
        )

        # Act
        output_path = self.service.generate_report(data, self.output_file)

        # Assert
        self.assertTrue(os.path.exists(output_path))

        # Verify content
        wb = load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws['A1'].value, "VIDEO PLAN: TEST PLAN")

        # Row 1 (header starts at row 4, data at 5)
        self.assertEqual(ws.cell(row=5, column=1).value, 1) # Scene
        # We can't easily check for images in openpyxl, but we can check if it didn't crash
        # And we check if B5 is empty (it should contain the image anchor, but value is None)
        self.assertIsNone(ws.cell(row=5, column=2).value)

        # Row 2 (Missing image)
        self.assertEqual(ws.cell(row=6, column=1).value, 2)
        self.assertEqual(ws.cell(row=6, column=2).value, "[No Image]")

if __name__ == "__main__":
    unittest.main()

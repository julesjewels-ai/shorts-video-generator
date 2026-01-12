"""Service for generating reports from processing results.

This service implements the Visual Data Export feature, decoupling report generation
from the batch orchestration logic.
"""

import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dance_loop_gen.core.models import ProcessingResult
from dance_loop_gen.utils.logger import setup_logger

logger = setup_logger()


class ReportService:
    """Service to generate detailed Excel reports."""

    def __init__(self):
        """Initialize the ReportService."""
        pass

    def generate_excel_report(self, results: List[ProcessingResult], output_dir: str) -> str:
        """Generate an Excel report from the processing results.

        Args:
            results: List of ProcessingResult objects.
            output_dir: Directory to save the report.

        Returns:
            Path to the generated Excel file.
        """
        logger.info(f"Generating Excel report for {len(results)} results...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Generation Report"

        # Define headers
        headers = [
            "Row Index", "Style", "Status", "Duration (s)",
            "Output Directory", "Error Message", "Timestamp"
        ]

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        success_font = Font(color="006100")

        failure_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        failure_font = Font(color="9C0006")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Write data
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for row_idx, result in enumerate(results, 2):
            # Row data
            data = [
                result.row_index,
                result.style,
                result.status,
                f"{result.duration_seconds:.2f}",
                result.output_dir or "N/A",
                result.error_message or "",
                current_time
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = left_align

                # Apply conditional formatting for Status column
                if col_idx == 3:  # Status column
                    cell.alignment = center_align
                    if value == "Success":
                        cell.fill = success_fill
                        cell.font = success_font
                    elif value == "Failed":
                        cell.fill = failure_fill
                        cell.font = failure_font

                # Hyperlink for Output Directory
                if col_idx == 5 and result.output_dir:
                    cell.hyperlink = result.output_dir
                    cell.style = "Hyperlink"

        # Adjust column widths
        for col_idx, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20  # Default width

        # Specific widths
        ws.column_dimensions['B'].width = 30  # Style
        ws.column_dimensions['E'].width = 50  # Output Dir
        ws.column_dimensions['F'].width = 50  # Error Message

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"batch_report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        wb.save(filepath)
        logger.info(f"Report saved to {filepath}")

        return filepath
